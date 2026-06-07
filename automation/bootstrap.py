"""
bootstrap.py — Migra o AvulsasRafael.xlsx local para a planilha Google Sheets.

Uso:
  python bootstrap.py --config config.json

O que faz:
  1. Abre a Sheets configurada (criar abas Hub/Demandas/Finalizadas se faltarem).
  2. Lê o xlsx local (sheet 'Demandas Design', 13 colunas no formato antigo).
  3. Re-mapeia colunas para o novo formato (insere HoraEntrega na posição C).
  4. Escreve tudo em 'Demandas' (preservando dados manuais).
  5. Inicializa o Hub com métricas básicas.

Roda UMA VEZ só. Depois é o sync.py.
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime
from typing import Optional

from openpyxl import load_workbook

import lib_sheets as ls
import lib_transform as lt
from sync import nomefinal_formula_for_row, apply_base_formatting, update_hub


def parse_hyperlink_cell(value) -> tuple[str, str]:
    """Recebe o valor de uma célula (str da fórmula HYPERLINK ou texto puro).
    Retorna (url, label)."""
    if not value:
        return "", ""
    s = str(value).strip()
    import re
    m = re.search(r'HYPERLINK\("([^"]*)",\s*"([^"]*)"\)', s, re.IGNORECASE)
    if m:
        return m.group(1), m.group(2)
    return s, s


def migrate(cfg: ls.Config) -> dict:
    if not cfg.source_xlsx_for_bootstrap:
        raise SystemExit("config.source_xlsx_for_bootstrap é obrigatório no bootstrap")

    # IMPORTANTE: abrir o spreadsheet ANTES de processar as linhas, pra que FORMULA_SEP
    # esteja correto quando rebuild_hyperlink for chamado.
    print("[bootstrap] Conectando ao Google Sheets...", file=sys.stderr)
    ss = ls.open_spreadsheet(cfg)

    print(f"[bootstrap] Abrindo xlsx: {cfg.source_xlsx_for_bootstrap}", file=sys.stderr)
    wb = load_workbook(cfg.source_xlsx_for_bootstrap, data_only=False)
    if "Demandas Design" not in wb.sheetnames:
        raise SystemExit("Sheet 'Demandas Design' não encontrada no xlsx")
    src = wb["Demandas Design"]

    # Headers do xlsx antigo (esperado):
    # ['Vertical','Entrega','TaskPai','TaskFilha','copy','Demandante','nomeCurto',
    #  'nomefinal','laminas','satatus','entregue','linkDrive','novo']
    headers_old = [c.value for c in src[1]]
    print(f"[bootstrap] Headers detectados: {headers_old}", file=sys.stderr)

    # Mapear nome → índice (1-based no openpyxl)
    def col_idx(name_options: list[str]) -> Optional[int]:
        for i, h in enumerate(headers_old, start=1):
            if h and str(h).strip().lower() in [n.lower() for n in name_options]:
                return i
        return None

    idx = {
        "Vertical": col_idx(["Vertical"]),
        "Entrega": col_idx(["Entrega"]),
        "TaskPai": col_idx(["TaskPai"]),
        "TaskFilha": col_idx(["TaskFilha"]),
        "copy": col_idx(["copy"]),
        "Demandante": col_idx(["Demandante"]),
        "nomeCurto": col_idx(["nomeCurto"]),
        "laminas": col_idx(["laminas"]),
        "status": col_idx(["status", "satatus"]),
        "entregue": col_idx(["entregue"]),
        "linkDrive": col_idx(["linkDrive"]),
    }
    missing = [k for k, v in idx.items() if v is None]
    if missing:
        raise SystemExit(f"Colunas faltando no xlsx: {missing}")

    # Coletar linhas
    migrated_rows = []
    for row in src.iter_rows(min_row=2, values_only=False):
        # Pular linhas vazias
        if all(c.value in (None, "") for c in row):
            continue

        get = lambda name: row[idx[name] - 1].value
        get_fmt = lambda name: row[idx[name] - 1].number_format

        entrega_val = get("Entrega")
        # Pode ser datetime ou string
        if isinstance(entrega_val, datetime):
            entrega_iso = entrega_val.date().isoformat()
        elif isinstance(entrega_val, date):
            entrega_iso = entrega_val.isoformat()
        elif entrega_val:
            entrega_iso = str(entrega_val)
        else:
            entrega_iso = ""

        # HoraEntrega: o xlsx antigo não tem; deixar vazio (sync futuro irá preencher)
        hora_entrega = ""

        # Converter fórmulas HYPERLINK do xlsx em rich-text link cells (clicáveis).
        def to_link_cell(v):
            if not v:
                return ""
            parsed = ls.parse_hyperlink_formula(str(v))
            if parsed:
                url, label = parsed
                return ls.make_link_cell(label, url)
            return str(v)

        def to_copy_cell(v):
            """Para a coluna copy: HYPERLINK→link azul; 'copyDescJira'→fallback cinza."""
            if not v:
                return ""
            sv = str(v).strip()
            parsed = ls.parse_hyperlink_formula(sv)
            if parsed:
                url, label = parsed
                return ls.make_link_cell(label, url)
            if sv.lower() in ("copydescjira", "copy desc jira"):
                return ls.make_fallback_copy_cell()
            return sv

        task_pai_val = to_link_cell(get("TaskPai"))
        task_filha_val = to_link_cell(get("TaskFilha"))
        copy_val = to_copy_cell(get("copy"))

        new_row = [""] * ls.N_COLS
        new_row[ls.COL_INDEX["Vertical"]] = get("Vertical") or ""
        new_row[ls.COL_INDEX["Entrega"]] = entrega_iso
        new_row[ls.COL_INDEX["HoraEntrega"]] = hora_entrega
        new_row[ls.COL_INDEX["TaskPai"]] = task_pai_val
        new_row[ls.COL_INDEX["TaskFilha"]] = task_filha_val
        new_row[ls.COL_INDEX["copy"]] = copy_val
        new_row[ls.COL_INDEX["Demandante"]] = get("Demandante") or ""
        new_row[ls.COL_INDEX["nomeCurto"]] = get("nomeCurto") or ""
        new_row[ls.COL_INDEX["nomefinal"]] = ""  # Vamos preencher com fórmula depois
        new_row[ls.COL_INDEX["status"]] = (get("status") or "").strip() if isinstance(get("status"), str) else (get("status") or "")
        new_row[ls.COL_INDEX["laminas"]] = get("laminas") or ""
        entregue_v = get("entregue")
        if isinstance(entregue_v, datetime):
            new_row[ls.COL_INDEX["entregue"]] = entregue_v.date().isoformat()
        elif isinstance(entregue_v, date):
            new_row[ls.COL_INDEX["entregue"]] = entregue_v.isoformat()
        else:
            new_row[ls.COL_INDEX["entregue"]] = entregue_v or ""
        new_row[ls.COL_INDEX["linkDrive"]] = get("linkDrive") or ""

        migrated_rows.append(new_row)

    print(f"[bootstrap] {len(migrated_rows)} linhas migradas do xlsx", file=sys.stderr)

    # ss já foi aberto no início (pra FORMULA_SEP estar setado durante o rebuild de fórmulas)

    # Criar abas
    ws_hub = ls.get_or_create_tab(ss, cfg.tab_hub, rows=80, cols=12)
    ws_d = ls.get_or_create_tab(ss, cfg.tab_demandas, rows=200, cols=ls.N_COLS)
    ws_f = ls.get_or_create_tab(ss, cfg.tab_finalizadas, rows=500, cols=ls.N_COLS)

    # Headers
    ls.ensure_header(ws_d)
    ls.ensure_header(ws_f)

    # Separar Feitos pra Finalizadas
    feitos = []
    ativas = []
    for r in migrated_rows:
        s = str(r[ls.COL_INDEX["status"]] or "").strip().lower()
        if s in {"feito", "concluido", "concluído", "done", "entregue"}:
            feitos.append(r)
        else:
            ativas.append(r)

    # Limpar valores E formatação do corpo (pra não herdar estilos do run anterior)
    def reset_body(ws):
        last_col_idx = ls.N_COLS - 1
        body = {
            "requests": [{
                "updateCells": {
                    "range": {
                        "sheetId": ws.id,
                        "startRowIndex": 1,
                        "endRowIndex": ws.row_count,
                        "startColumnIndex": 0,
                        "endColumnIndex": ls.N_COLS,
                    },
                    "fields": "userEnteredValue,userEnteredFormat,textFormatRuns",
                }
            }]
        }
        ws.spreadsheet.batch_update(body)

    reset_body(ws_d)
    reset_body(ws_f)

    # Escrever Demandas (com nomefinal por linha) — usando rich-text link nativo
    if ativas:
        rows_to_write = []
        for i, row in enumerate(ativas, start=2):
            row = list(row)
            row[ls.COL_INDEX["nomefinal"]] = nomefinal_formula_for_row(i)
            rows_to_write.append(row)
        ls.write_rows_native(ws_d, 2, rows_to_write)
        print(f"[bootstrap] {len(ativas)} linhas escritas em '{cfg.tab_demandas}'", file=sys.stderr)

    if feitos:
        rows_to_write = []
        for i, row in enumerate(feitos, start=2):
            row = list(row)
            row[ls.COL_INDEX["nomefinal"]] = nomefinal_formula_for_row(i)
            rows_to_write.append(row)
        ls.write_rows_native(ws_f, 2, rows_to_write)
        print(f"[bootstrap] {len(feitos)} linhas escritas em '{cfg.tab_finalizadas}'", file=sys.stderr)

    # Formatação
    apply_base_formatting(ws_d)
    apply_base_formatting(ws_f)

    # Hub inicial
    update_hub(ss, cfg, run_date_iso=date.today().isoformat())

    return {
        "migrated_total": len(migrated_rows),
        "ativas": len(ativas),
        "finalizadas": len(feitos),
        "spreadsheet_url": f"https://docs.google.com/spreadsheets/d/{cfg.spreadsheet_id}",
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", default="config.json")
    args = parser.parse_args()

    cfg = ls.Config.load(args.config)
    result = migrate(cfg)
    print(json.dumps(result, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
